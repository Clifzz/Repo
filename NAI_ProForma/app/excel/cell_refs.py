from openpyxl.utils import get_column_letter

BLDG_ROW = {
    "building_name": 1, "start_year": 2, "start_month": 3, "years": 4,
    "total_sqft": 5, "occupied_sqft": 6, "opex_psf": 7, "opex_growth": 8,
    "market_avg_rate": 9, "market_growth": 10, "cap_rate": 11,
    "cap_delta": 12, "weighted_avg_rate": 13,
}
BLDG_VAL_COL = 2

TENANT_HEADER_ROW = 17
TENANT_DATA_START_ROW = 18
TENANT_COL = {
    "name": 1, "suite": 2, "sqft": 3, "rate_psf": 4,
    "year1_override": 5, "lease_exp": 6, "projection_type": 7,
    "growth_rate": 8, "flat_increase": 9, "pct_increase": 10,
    "renewed": 11, "renewal_start": 12, "renewal_term_years": 13,
    "renewal_projection_type": 14, "renewal_growth_rate": 15,
    "renewal_flat_increase": 16, "renewal_pct_increase": 17,
}

PF_TITLE_ROW = 1
PF_EXPIRING_SF_ROW = 4
PF_EXPIRING_PCT_ROW = 5
PF_HEADER_ROW = 6
PF_TENANT_SUBHDR_ROW = 7
PF_TENANT_DATA_START = 8
PF_ASSUMP_START_COL = 1
PF_ASSUMP_END_COL = 7
PF_PROJ_START_COL = 9


def bldg_ref(param: str) -> str:
    return f"Inputs!{get_column_letter(BLDG_VAL_COL)}{BLDG_ROW[param]}"


def tenant_ref(tenant_idx: int, col_name: str) -> str:
    row = TENANT_DATA_START_ROW + tenant_idx
    col = get_column_letter(TENANT_COL[col_name])
    return f"Inputs!{col}{row}"


def last_tenant_row(num_tenants: int) -> int:
    return TENANT_DATA_START_ROW + num_tenants - 1
