from __future__ import annotations
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models.session import ProFormaSession
from app.excel.cell_refs import (
    bldg_ref, tenant_ref, last_tenant_row,
    TENANT_DATA_START_ROW, TENANT_COL,
    PF_TITLE_ROW, PF_EXPIRING_SF_ROW, PF_EXPIRING_PCT_ROW,
    PF_HEADER_ROW, PF_TENANT_SUBHDR_ROW, PF_TENANT_DATA_START,
    PF_ASSUMP_START_COL, PF_ASSUMP_END_COL, PF_PROJ_START_COL,
)

_THIN = Side("thin")
_CB   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FILL  = PatternFill("solid", fgColor="4F81BD")
_SUBHDR_FILL = PatternFill("solid", fgColor="4A4A4A")
_LGT_GREEN = PatternFill("solid", fgColor="C6EFCE")
_DRK_GREEN = PatternFill("solid", fgColor="00B050")
_RED_FONT  = Font(bold=True, color="C8102E")
_BOLD      = Font(bold=True)
_BOLD_WH   = Font(bold=True, color="FFFFFF")


def _months_old(exp_ref: str) -> str:
    return (f"IF(MONTH({exp_ref})=1,1,"
            f"MIN(MONTH({exp_ref})+IF(DAY({exp_ref})>=15,1,0),12))")


def _base_rate_expr(y1ov: str, sf: str, rate: str) -> str:
    return f"IF({y1ov}<>\"\",{y1ov}/{sf},{rate})"


def _rent_y1(y1ov: str, sf: str, rate: str) -> str:
    return f"=IF({y1ov}<>\"\",{y1ov},{sf}*{rate})"


def _rent_compounded(n: int, y1ov: str, sf: str, rate: str, grow: str) -> str:
    base = f"IF({y1ov}<>\"\",{y1ov},{sf}*{rate})"
    return f"={base}*(1+{grow})^{n}"


def _rent_prorated_flat(n: int, y1ov: str, sf: str, rate: str, flat: str, exp: str) -> str:
    br  = _base_rate_expr(y1ov, sf, rate)
    mo  = _months_old(exp)
    mn  = f"MAX(12-({mo}),0)"
    old = f"(({br})+{flat}*{n - 1})"
    new = f"(({br})+{flat}*{n})"
    return f"={sf}*(({old}*({mo})+{new}*{mn})/12)"


def _rent_prorated_pct(n: int, y1ov: str, sf: str, rate: str, pct: str, exp: str) -> str:
    br  = _base_rate_expr(y1ov, sf, rate)
    mo  = _months_old(exp)
    mn  = f"MAX(12-({mo}),0)"
    old = f"(({br})*(1+{pct})^{n - 1})"
    new = f"(({br})*(1+{pct})^{n})"
    return f"={sf}*(({old}*({mo})+{new}*{mn})/12)"


def _tenant_formula(t_idx: int, y_idx: int, session: ProFormaSession) -> str:
    ten  = session.tenants[t_idx]
    sf   = tenant_ref(t_idx, "sqft")
    rate = tenant_ref(t_idx, "rate_psf")
    y1ov = tenant_ref(t_idx, "year1_override")
    grow = tenant_ref(t_idx, "growth_rate")
    flat = tenant_ref(t_idx, "flat_increase")
    pct  = tenant_ref(t_idx, "pct_increase")
    exp  = tenant_ref(t_idx, "lease_exp")

    if y_idx == 0:
        formula = _rent_y1(y1ov, sf, rate)
    elif ten.projection_type == "compounded":
        formula = _rent_compounded(y_idx, y1ov, sf, rate, grow)
    elif ten.flat_increase > 0:
        formula = _rent_prorated_flat(y_idx, y1ov, sf, rate, flat, exp)
    else:
        # pct_increase=0 also works correctly here: (1+0)^n = 1, formula holds flat rate
        formula = _rent_prorated_pct(y_idx, y1ov, sf, rate, pct, exp)

    if ten.renewed and ten.renewal_start and y_idx > 0:
        yr       = session.start_year + y_idx
        rs_ref   = tenant_ref(t_idx, "renewal_start")
        rg_ref   = tenant_ref(t_idx, "renewal_growth_rate")
        rf_ref   = tenant_ref(t_idx, "renewal_flat_increase")
        rp_ref   = tenant_ref(t_idx, "renewal_pct_increase")
        rtype    = ten.renewal_projection_type
        ren_yr   = int(ten.renewal_start[-4:])
        n_ren    = yr - ren_yr

        ren_base = _base_rate_expr(y1ov, sf, rate)

        if rtype == "compounded":
            ren = f"{sf}*({ren_base})*(1+{rg_ref})^{n_ren}"
        elif ten.renewal_flat_increase > 0:
            ren = f"{sf}*(({ren_base})+{rf_ref}*{n_ren})"
        else:
            # renewal_pct_increase=0 also works correctly: (1+0)^n = 1, formula holds flat rate
            ren = f"{sf}*(({ren_base})*(1+{rp_ref})^{n_ren})"

        formula = f"=IF(DATE({yr},1,1)>={rs_ref},{ren},{formula[1:]})"

    return formula


def write_proforma_sheet(ws, session: ProFormaSession) -> None:
    years = session.years; num_t = len(session.tenants)
    sc = PF_ASSUMP_START_COL; ps = PF_PROJ_START_COL
    proj_end = ps + years - 1

    # Title
    ws.merge_cells(start_row=PF_TITLE_ROW, start_column=sc,
                   end_row=PF_TITLE_ROW + 2, end_column=proj_end)
    tc = ws.cell(row=PF_TITLE_ROW, column=sc,
                 value=f"={bldg_ref('building_name')}&\" Pro-Forma / Rent Roll\"")
    tc.font = Font(size=18, bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor="333333")
    tc.alignment = Alignment(horizontal="center", vertical="center")

    # Expiring SF/% rows
    exp_col = get_column_letter(TENANT_COL["lease_exp"])
    sf_col  = get_column_letter(TENANT_COL["sqft"])
    t0 = TENANT_DATA_START_ROW; tN = last_tenant_row(num_t)
    ws.cell(row=PF_EXPIRING_SF_ROW, column=sc, value="SF Expiring").font = _BOLD
    ws.cell(row=PF_EXPIRING_PCT_ROW, column=sc, value="% Expiring").font = _BOLD
    for i in range(years):
        yr = session.start_year + i; col = ps + i; cl = get_column_letter(col)
        rng_e = f"Inputs!{exp_col}{t0}:Inputs!{exp_col}{tN}"
        rng_s = f"Inputs!{sf_col}{t0}:Inputs!{sf_col}{tN}"
        sf_f  = (f"=SUMIF({rng_e},\">=\"&DATE({yr},1,1),{rng_s})"
                 f"-SUMIF({rng_e},\">=\"&DATE({yr+1},1,1),{rng_s})")
        c1 = ws.cell(row=PF_EXPIRING_SF_ROW, column=col, value=sf_f)
        c1.border = _CB; c1.alignment = Alignment(horizontal="center")
        c2 = ws.cell(row=PF_EXPIRING_PCT_ROW, column=col,
                     value=f"={cl}{PF_EXPIRING_SF_ROW}/{bldg_ref('total_sqft')}")
        c2.number_format = "0.00%"; c2.border = _CB
        c2.alignment = Alignment(horizontal="center")

    # Column headers
    ws.merge_cells(start_row=PF_HEADER_ROW, start_column=sc,
                   end_row=PF_HEADER_ROW, end_column=PF_ASSUMP_END_COL)
    h = ws.cell(row=PF_HEADER_ROW, column=sc, value="ASSUMPTIONS & TENANCY DETAIL")
    h.font = _BOLD_WH; h.fill = _HDR_FILL; h.alignment = Alignment(horizontal="center")
    for i in range(years):
        c = ws.cell(row=PF_HEADER_ROW, column=ps + i, value=str(session.start_year + i))
        c.font = _BOLD_WH; c.fill = _HDR_FILL; c.alignment = Alignment(horizontal="center"); c.border = _CB

    for j, lbl in enumerate(["TENANCY","SUITE #","SIZE (SF)","$/SF",
                              "LEASE EXP","TERM REM.","AS IS RENT"], start=sc):
        c = ws.cell(row=PF_TENANT_SUBHDR_ROW, column=j, value=lbl)
        c.font = _BOLD_WH; c.fill = _SUBHDR_FILL; c.alignment = Alignment(horizontal="center"); c.border = _CB

    # Tenant rows
    for t_idx in range(num_t):
        r  = PF_TENANT_DATA_START + t_idx; tr = TENANT_DATA_START_ROW + t_idx
        sf = tenant_ref(t_idx, "sqft"); rt = tenant_ref(t_idx, "rate_psf")
        ex = tenant_ref(t_idx, "lease_exp")
        as_of = (f"DATE({bldg_ref('start_year')},{bldg_ref('start_month')},1)")
        term_f = (f"=DATEDIF({as_of},{ex},\"Y\")&\" yrs, \""
                  f"&DATEDIF({as_of},{ex},\"YM\")&\" mos\"")

        for col, val, fmt in [
            (1, f"=Inputs!A{tr}", None), (2, f"=Inputs!B{tr}", None),
            (3, f"={sf}", "#,##0"), (4, f"={rt}", '"$"#,##0.00'),
            (5, f"={ex}", "MM-DD-YYYY"), (6, term_f, None),
            (7, f"={sf}*{rt}", '"$"#,##0'),
        ]:
            c = ws.cell(row=r, column=col, value=val)
            if fmt: c.number_format = fmt
            c.border = _CB

        for y_idx in range(years):
            formula = _tenant_formula(t_idx, y_idx, session)
            c = ws.cell(row=r, column=ps + y_idx, value=formula)
            c.number_format = '"$"#,##0'; c.border = _CB
            c.alignment = Alignment(horizontal="center")

    # Roll-up rows
    rr = PF_TENANT_DATA_START + num_t + 1
    rent_cells = []; exp_cells = []; opex_cells = []; noi_cells = []

    ws.cell(row=rr, column=sc, value="Rental Revenue").font = _BOLD
    for i in range(years):
        col = ps + i; cl = get_column_letter(col)
        t_range = f"{cl}{PF_TENANT_DATA_START}:{cl}{PF_TENANT_DATA_START + num_t - 1}"
        c = ws.cell(row=rr, column=col, value=f"=SUM({t_range})")
        c.number_format = '"$"#,##0'; c.border = _CB; rent_cells.append(f"{cl}{rr}")
    rr += 1

    ws.cell(row=rr, column=sc, value="Expense Revenue").font = _BOLD
    for i in range(years):
        col = ps + i; cl = get_column_letter(col)
        f_val = (f"={bldg_ref('total_sqft')}*{bldg_ref('opex_psf')}"
                 if i == 0
                 else f"={get_column_letter(ps+i-1)}{rr}*(1+{bldg_ref('opex_growth')})")
        c = ws.cell(row=rr, column=col, value=f_val)
        c.number_format = '"$"#,##0'; c.border = _CB; exp_cells.append(f"{cl}{rr}")
    rr += 1

    ws.cell(row=rr, column=sc, value="Total Gross Revenue").font = _BOLD
    for i in range(years):
        c = ws.cell(row=rr, column=ps+i, value=f"={rent_cells[i]}+{exp_cells[i]}")
        c.number_format = '"$"#,##0'; c.border = _CB
    rr += 1

    ws.cell(row=rr, column=sc, value="Operating Expenses").font = _RED_FONT
    for i in range(years):
        col = ps + i; cl = get_column_letter(col)
        f_v = (f"={bldg_ref('total_sqft')}*{bldg_ref('opex_psf')}"
               f"*(1+{bldg_ref('opex_growth')})^{i}")
        c = ws.cell(row=rr, column=col, value=f_v)
        c.number_format = '"$"#,##0'; c.font = _RED_FONT; c.border = _CB; opex_cells.append(f"{cl}{rr}")
    rr += 1

    ws.cell(row=rr, column=sc, value="OpEx $/SF").font = _RED_FONT
    for i in range(years):
        c = ws.cell(row=rr, column=ps+i,
                    value=f"={opex_cells[i]}/{bldg_ref('total_sqft')}")
        c.number_format = '"$"#,##0.00'; c.font = _RED_FONT; c.border = _CB
    rr += 1

    ws.cell(row=rr, column=sc, value="Net Operating Income").font = _BOLD
    for i in range(years):
        col = ps + i; cl = get_column_letter(col)
        c = ws.cell(row=rr, column=col,
                    value=f"={rent_cells[i]}+{exp_cells[i]}-{opex_cells[i]}")
        c.number_format = '"$"#,##0'; c.border = _CB; noi_cells.append(f"{cl}{rr}")
    rr += 2

    ws.cell(row=rr, column=sc, value="Market Avg Rate").font = _BOLD
    for i in range(years):
        c = ws.cell(row=rr, column=ps+i,
                    value=f"={bldg_ref('market_avg_rate')}*(1+{bldg_ref('market_growth')})^{i}")
        c.number_format = '"$"#,##0.00'; c.border = _CB
    rr += 1

    ws.cell(row=rr, column=sc, value="Weighted Avg Rate").font = _BOLD
    for i in range(years):
        c = ws.cell(row=rr, column=ps+i,
                    value=f"={bldg_ref('weighted_avg_rate')}*(1+{bldg_ref('market_growth')})^{i}")
        c.number_format = '"$"#,##0.00'; c.border = _CB
    rr += 2

    ws.cell(row=rr, column=sc, value="Building Value").font = _BOLD
    val_cells = []
    for i in range(years):
        col = ps + i; cl = get_column_letter(col)
        c = ws.cell(row=rr, column=col, value=f"={noi_cells[i]}/{bldg_ref('cap_rate')}")
        c.number_format = '"$"#,##0'; c.border = _CB; val_cells.append(f"{cl}{rr}")
    rr += 1

    ws.cell(row=rr, column=sc, value="Value / SF").font = _BOLD
    for i in range(years):
        c = ws.cell(row=rr, column=ps+i,
                    value=f"={val_cells[i]}/{bldg_ref('total_sqft')}")
        c.number_format = '"$"#,##0.00'; c.border = _CB
    rr += 2

    # Cap rate sensitivity
    noi_y1 = noi_cells[0]
    ws.cell(row=rr, column=sc, value="CAP RATE SENSITIVITY").font = _BOLD; rr += 1
    for j, lbl in enumerate(["Cap Rate","NOI","Value","$/SF"], start=sc):
        c = ws.cell(row=rr, column=j, value=lbl)
        c.font = _BOLD; c.border = _CB; c.alignment = Alignment(horizontal="center")
    rr += 1

    cr = bldg_ref("cap_rate"); cd = bldg_ref("cap_delta")
    for fill, expr in [
        (_LGT_GREEN, f"{cr}-{cd}"),
        (_DRK_GREEN, cr),
        (_LGT_GREEN, f"{cr}+{cd}"),
    ]:
        for j, (val, fmt) in enumerate(zip(
            [f"=({expr})", f"={noi_y1}", f"={noi_y1}/({expr})",
             f"={noi_y1}/({expr})/{bldg_ref('total_sqft')}"],
            ["0.00%", '"$"#,##0', '"$"#,##0', '"$"#,##0.00'],
        ), start=sc):
            c = ws.cell(row=rr, column=j, value=val)
            c.number_format = fmt; c.fill = fill; c.border = _CB
            c.alignment = Alignment(horizontal="center")
        rr += 1

    for col in ws.columns:
        col_idx = col[0].column
        max_len = max((len(str(c.value)) if c.value and not str(c.value).startswith("=") else 0) for c in col)
        min_w = 14 if col_idx >= PF_PROJ_START_COL else 4
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min(max_len + 4, 40), min_w)
