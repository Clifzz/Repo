from __future__ import annotations
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.models.session import ProFormaSession
from app.excel.cell_refs import (
    BLDG_ROW, BLDG_VAL_COL, TENANT_HEADER_ROW,
    TENANT_DATA_START_ROW, TENANT_COL,
)

_LABEL_FILL  = PatternFill("solid", fgColor="F0F0F0")
_HDR_FILL    = PatternFill("solid", fgColor="C8102E")
_ALT_FILL    = PatternFill("solid", fgColor="F9F9F9")
_BOLD        = Font(bold=True)
_BOLD_WHITE  = Font(bold=True, color="FFFFFF")

_BLDG_PARAMS = [
    ("Building Name",    "building_name"),
    ("Start Year",       "start_year"),
    ("Start Month",      "start_month"),
    ("Years to Project", "years"),
    ("Total SF",         "total_sqft"),
    ("Occupied SF",      "occupied_sqft"),
    ("OpEx / SF",        "opex_psf"),
    ("OpEx Growth Rate", "opex_growth"),
    ("Market Avg Rate",  "market_avg_rate"),
    ("Market Growth %",  "market_growth"),
    ("Cap Rate",         "cap_rate"),
    ("Cap Delta",        "cap_delta"),
    ("Weighted Avg Rate","weighted_avg_rate"),
]

_BLDG_VALUES = {
    "building_name":   lambda s: s.building_name,
    "start_year":      lambda s: s.start_year,
    "start_month":     lambda s: s.start_month,
    "years":           lambda s: s.years,
    "total_sqft":      lambda s: s.total_sqft,
    "occupied_sqft":   lambda s: s.occupied_sqft,
    "opex_psf":        lambda s: s.opex_psf,
    "opex_growth":     lambda s: 0.025,
    "market_avg_rate": lambda s: s.market_avg_rate,
    "market_growth":   lambda s: s.market_growth_pct,
    "cap_rate":        lambda s: s.cap_rate,
    "cap_delta":       lambda s: s.cap_delta,
}


def _insert_logo(ws, logo_png: str | None, cell: str, width: int = 214, height: int = 47) -> None:
    if not logo_png:
        return
    try:
        from openpyxl.drawing.image import Image as XlImg
        img = XlImg(logo_png)
        img.width = width
        img.height = height
        ws.add_image(img, cell)
    except Exception:
        pass


def write_inputs_sheet(
    ws,
    session: ProFormaSession,
    irr_data: dict | None = None,
    logo_png: str | None = None,
) -> None:
    ws.title = "Inputs"
    num_t = len(session.tenants)
    first_t = TENANT_DATA_START_ROW
    last_t = TENANT_DATA_START_ROW + num_t - 1
    sf_col = get_column_letter(TENANT_COL["sqft"])
    rt_col = get_column_letter(TENANT_COL["rate_psf"])
    max_tenant_col = get_column_letter(max(TENANT_COL.values()))

    # ── Building parameter rows (rows 1-13, cols A-B) ────────────────────
    for label, key in _BLDG_PARAMS:
        row = BLDG_ROW[key]
        lbl_cell = ws.cell(row=row, column=1, value=label)
        lbl_cell.font = _BOLD
        lbl_cell.fill = _LABEL_FILL
        lbl_cell.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 16
        if key == "weighted_avg_rate":
            if num_t > 0:
                sf_r = f"{sf_col}{first_t}:{sf_col}{last_t}"
                rt_r = f"{rt_col}{first_t}:{rt_col}{last_t}"
                ws.cell(row=row, column=BLDG_VAL_COL,
                        value=f"=SUMPRODUCT({sf_r},{rt_r})/SUM({sf_r})")
        else:
            ws.cell(row=row, column=BLDG_VAL_COL, value=_BLDG_VALUES[key](session))

    # ── Logo + title block (right-hand empty area, cols D onward, rows 1-8) ──
    # Building params only occupy cols A-B in rows 1-13; cols D+ are empty here.
    _insert_logo(ws, logo_png, "D1", width=214, height=47)

    ws.merge_cells("D4:K4")
    t = ws.cell(row=4, column=4, value="NAI Pro Forma Generator — Inputs")
    t.font = Font(size=13, bold=True, color="C8102E")
    t.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("D5:K5")
    b = ws.cell(row=5, column=4, value=session.building_name or "")
    b.font = Font(size=11, bold=True, color="3A3A3C")
    b.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("D6:K6")
    d = ws.cell(row=6, column=4,
                value=f"Generated: {datetime.now().strftime('%B %d, %Y')}")
    d.font = Font(size=9, color="8E8E93")
    d.alignment = Alignment(horizontal="left", vertical="center")

    # ── Gap rows: "TENANT SCHEDULE" section header ────────────────────────
    # Rows 14-16 are empty (gap between building params ending row 13 and
    # tenant header row 17). Use row 15 for the section header.
    ws.row_dimensions[14].height = 8
    ws.merge_cells(f"A15:{max_tenant_col}15")
    sec = ws.cell(row=15, column=1, value="TENANT SCHEDULE")
    sec.font = _BOLD_WHITE
    sec.fill = _HDR_FILL
    sec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[15].height = 18
    ws.row_dimensions[16].height = 4

    # ── Tenant column headers (row 17) ────────────────────────────────────
    for col_name, col_idx in TENANT_COL.items():
        c = ws.cell(row=TENANT_HEADER_ROW, column=col_idx,
                    value=col_name.replace("_", " ").title())
        c.font = _BOLD_WHITE
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[TENANT_HEADER_ROW].height = 16

    # ── Tenant data rows ──────────────────────────────────────────────────
    for i, tenant in enumerate(session.tenants):
        row = TENANT_DATA_START_ROW + i
        values = {
            "name": tenant.name, "suite": tenant.suite,
            "sqft": tenant.sqft, "rate_psf": tenant.rate_psf,
            "year1_override": tenant.year1_override if tenant.year1_override is not None else "",
            "lease_exp": datetime.strptime(tenant.lease_exp, "%m-%d-%Y") if tenant.lease_exp and tenant.lease_exp.strip() else "",
            "projection_type": tenant.projection_type,
            "growth_rate": tenant.growth_rate, "flat_increase": tenant.flat_increase,
            "pct_increase": tenant.pct_increase, "renewed": tenant.renewed,
            "renewal_start": datetime.strptime(tenant.renewal_start, "%m-%d-%Y") if tenant.renewal_start and tenant.renewal_start.strip() else "",
            "renewal_term_years": tenant.renewal_term_years,
            "renewal_projection_type": tenant.renewal_projection_type,
            "renewal_growth_rate": tenant.renewal_growth_rate,
            "renewal_flat_increase": tenant.renewal_flat_increase,
            "renewal_pct_increase": tenant.renewal_pct_increase,
        }
        row_fill = _ALT_FILL if i % 2 == 1 else None
        for col_name, col_idx in TENANT_COL.items():
            cell = ws.cell(row=row, column=col_idx, value=values[col_name])
            if col_name in ("lease_exp", "renewal_start") and values[col_name]:
                cell.number_format = "MM-DD-YYYY"
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[row].height = 15

    # ── Return Analysis block ─────────────────────────────────────────────
    if irr_data is not None:
        irr_str = f"{irr_data['irr']:.2%}" if irr_data.get("irr") is not None else "N/A"
        start_row = TENANT_DATA_START_ROW + max(len(session.tenants), 1) + 2

        ws.merge_cells(f"A{start_row}:B{start_row}")
        hdr = ws.cell(row=start_row, column=1, value="Return Analysis")
        hdr.font = _BOLD_WHITE
        hdr.fill = _HDR_FILL
        hdr.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[start_row].height = 18

        rows_data = [
            ("Purchase Price", f"${irr_data['effective_purchase']:,.0f}"),
            ("Exit Cap Rate",  f"{irr_data['effective_exit_cap']:.2%}"),
            ("Exit Value",     f"${irr_data['exit_value']:,.0f}"),
            ("IRR",            irr_str),
            (f"NPV ({session.discount_rate:.1%} discount)", f"${irr_data['npv']:,.0f}"),
        ]
        for j, (label, value) in enumerate(rows_data, start=start_row + 1):
            lbl_cell = ws.cell(row=j, column=1, value=label)
            lbl_cell.font = _BOLD
            lbl_cell.fill = _LABEL_FILL
            lbl_cell.alignment = Alignment(vertical="center", indent=1)
            ws.cell(row=j, column=2, value=value)
            ws.row_dimensions[j].height = 15

    # ── Column widths ─────────────────────────────────────────────────────
    for col in ws.columns:
        col_idx = col[0].column
        max_len = max(
            (len(str(c.value)) if c.value and not str(c.value).startswith("=") else 0)
            for c in col
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 26)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 20)
