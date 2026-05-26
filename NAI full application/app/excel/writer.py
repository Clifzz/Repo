from __future__ import annotations
import os
import tempfile
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
from app.models.session import ProFormaSession
from app.excel.inputs_sheet import write_inputs_sheet
from app.excel.proforma_sheet import write_proforma_sheet

_LOGO_SVG = Path(__file__).parent.parent / "assets" / "nai_logo.svg"


def _make_logo_png(width: int = 428, height: int = 94) -> str | None:
    """Render nai_logo.svg to a temp PNG via PySide6; return path or None."""
    if not _LOGO_SVG.exists():
        return None
    try:
        from PySide6.QtSvg import QSvgRenderer
        from PySide6.QtGui import QImage, QPainter
        from PySide6.QtCore import Qt
        renderer = QSvgRenderer(str(_LOGO_SVG))
        img = QImage(width, height, QImage.Format.Format_ARGB32)
        img.fill(Qt.GlobalColor.white)
        painter = QPainter(img)
        renderer.render(painter)
        painter.end()
        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        tmp.close()
        img.save(tmp.name)
        return tmp.name
    except Exception:
        return None


def write_workbook(
    session: ProFormaSession,
    output_path: str,
    irr_data: dict | None = None,
) -> None:
    logo_png = _make_logo_png()
    try:
        wb = Workbook()
        write_inputs_sheet(wb.active, session, irr_data=irr_data, logo_png=logo_png)
        write_proforma_sheet(wb.create_sheet("ProForma"), session, logo_png=logo_png)
        _write_notes_sheet(wb.create_sheet("Notes"), session)
        wb.save(output_path)
    finally:
        if logo_png:
            try:
                os.unlink(logo_png)
            except OSError:
                pass


def _write_notes_sheet(ws, session: ProFormaSession) -> None:
    ws.title = "Notes"
    header = ws.cell(row=1, column=1, value="Deal Notes & Assumptions")
    header.font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 80
    if session.notes.strip():
        for i, line in enumerate(session.notes.splitlines(), start=2):
            ws.cell(row=i, column=1, value=line)
    else:
        ws.cell(row=2, column=1, value="No notes entered.")
