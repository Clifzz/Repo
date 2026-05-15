import os
from app.excel.writer import write_workbook
from openpyxl import load_workbook


def test_creates_two_sheets(basic_session, tmp_path):
    out = str(tmp_path / "test.xlsx")
    write_workbook(basic_session, out)
    wb = load_workbook(out)
    assert "Inputs" in wb.sheetnames
    assert "ProForma" in wb.sheetnames


def test_file_exists(basic_session, tmp_path):
    out = str(tmp_path / "test.xlsx")
    write_workbook(basic_session, out)
    assert os.path.exists(out)
