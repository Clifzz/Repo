# tests/test_inputs_sheet.py
import pytest
from openpyxl import Workbook
from app.excel.inputs_sheet import write_inputs_sheet
from app.excel.cell_refs import bldg_ref, tenant_ref, TENANT_DATA_START_ROW


def test_building_params_written(basic_session):
    wb = Workbook(); ws = wb.active
    write_inputs_sheet(ws, basic_session)
    assert ws["B11"].value == pytest.approx(basic_session.cap_rate)


def test_tenant_sqft_written(basic_session):
    wb = Workbook(); ws = wb.active
    write_inputs_sheet(ws, basic_session)
    assert ws[f"C{TENANT_DATA_START_ROW}"].value == pytest.approx(5000.0)


def test_weighted_avg_rate_is_formula(basic_session):
    wb = Workbook(); ws = wb.active
    write_inputs_sheet(ws, basic_session)
    assert str(ws["B13"].value).startswith("=")


def test_bldg_ref():
    assert bldg_ref("cap_rate") == "Inputs!B11"
    assert bldg_ref("total_sqft") == "Inputs!B5"


def test_tenant_ref():
    assert tenant_ref(0, "sqft") == "Inputs!C18"
    assert tenant_ref(1, "rate_psf") == "Inputs!D19"
