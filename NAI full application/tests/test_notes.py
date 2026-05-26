import json
import pytest
from app.models.session import ProFormaSession


def test_notes_default_empty():
    s = ProFormaSession()
    assert s.notes == ""


def test_notes_round_trip(basic_session):
    basic_session.notes = "Deal note line 1.\nLine 2."
    restored = ProFormaSession.from_json(basic_session.to_json())
    assert restored.notes == "Deal note line 1.\nLine 2."


def test_old_json_without_notes_loads_empty():
    old = {
        "building_name": "Old", "start_year": 2024, "start_month": 1,
        "years": 5, "total_sqft": 1000.0, "occupied_sqft": 800.0,
        "opex_psf": 5.0, "market_avg_rate": 20.0, "market_growth_pct": 0.02,
        "cap_rate": 0.065, "cap_delta": 0.0025, "tenants": [],
    }
    s = ProFormaSession.from_json(json.dumps(old))
    assert s.notes == ""
    assert s.purchase_price == 0.0
    assert s.exit_cap_rate == 0.0
    assert s.discount_rate == 0.08


def test_notes_preserved_in_template(basic_session):
    from app.db.database import init_db, save_template, get_template
    basic_session.notes = "Test assumption"
    conn = init_db(":memory:")
    tid = save_template("T", basic_session, conn=conn)
    row = get_template(tid, conn=conn)
    restored = ProFormaSession.from_json(row["inputs_json"])
    conn.close()
    assert restored.notes == "Test assumption"


def test_notes_page_bind_populates_editor(qtbot):
    from app.ui.wizard.page_notes import NotesPage
    s = ProFormaSession()
    s.notes = "Existing assumption"
    page = NotesPage(s)
    qtbot.addWidget(page)
    page.bind(s)
    assert page._edit.toPlainText() == "Existing assumption"


def test_pdf_contains_notes(basic_session):
    from app.logic.calculator import calculate_proforma
    from app.excel.pdf_writer import _build_html
    basic_session.notes = "Important assumption here."
    result = calculate_proforma(basic_session)
    html = _build_html(basic_session, result, irr_data=None)
    assert "Important assumption here." in html


def test_pdf_omits_notes_section_when_empty(basic_session):
    from app.logic.calculator import calculate_proforma
    from app.excel.pdf_writer import _build_html
    basic_session.notes = ""
    result = calculate_proforma(basic_session)
    html = _build_html(basic_session, result, irr_data=None)
    assert "Deal Notes" not in html


def test_excel_notes_sheet_exists(basic_session, tmp_path):
    from openpyxl import load_workbook
    from app.excel.writer import write_workbook
    basic_session.notes = "Test note."
    path = str(tmp_path / "test.xlsx")
    write_workbook(basic_session, path)
    wb = load_workbook(path)
    assert "Notes" in wb.sheetnames


def test_excel_notes_sheet_contains_text(basic_session, tmp_path):
    from openpyxl import load_workbook
    from app.excel.writer import write_workbook
    basic_session.notes = "My deal note."
    path = str(tmp_path / "test.xlsx")
    write_workbook(basic_session, path)
    wb = load_workbook(path)
    ws = wb["Notes"]
    values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any("My deal note." in str(v) for v in values if v)


def test_excel_notes_sheet_placeholder_when_empty(basic_session, tmp_path):
    from openpyxl import load_workbook
    from app.excel.writer import write_workbook
    basic_session.notes = ""
    path = str(tmp_path / "test.xlsx")
    write_workbook(basic_session, path)
    wb = load_workbook(path)
    ws = wb["Notes"]
    values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any("No notes" in str(v) for v in values if v)
