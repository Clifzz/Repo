import pytest
from app.logic.calculator import _irr, _npv, calculate_irr_npv


def test_irr_simple_known_value():
    # Invest 1000, receive 100/yr for 9 years, receive 1100 in year 10 → IRR ≈ 10%
    cf = [-1000] + [100] * 9 + [1100]
    assert _irr(cf) == pytest.approx(0.10, abs=1e-4)


def test_npv_at_zero_discount_equals_sum():
    cf = [-1000, 400, 400, 400]
    assert _npv(cf, 0.0) == pytest.approx(200.0)


def test_npv_positive_at_low_discount(basic_session):
    from app.logic.calculator import calculate_proforma
    result = calculate_proforma(basic_session)
    basic_session.purchase_price = result["values"][0]
    basic_session.discount_rate = 0.01
    data = calculate_irr_npv(basic_session, result)
    assert data["npv"] > 0


def test_irr_returns_none_for_no_valid_irr():
    # All-positive cash flows → no valid IRR, algorithm diverges
    assert _irr([100, 100, 100]) is None


def test_calculate_irr_npv_uses_y1_value_when_purchase_zero(basic_session):
    from app.logic.calculator import calculate_proforma
    result = calculate_proforma(basic_session)
    basic_session.purchase_price = 0.0
    data = calculate_irr_npv(basic_session, result)
    assert data["effective_purchase"] == pytest.approx(result["values"][0])


def test_calculate_irr_npv_uses_explicit_purchase(basic_session):
    from app.logic.calculator import calculate_proforma
    result = calculate_proforma(basic_session)
    basic_session.purchase_price = 999_999.0
    data = calculate_irr_npv(basic_session, result)
    assert data["effective_purchase"] == pytest.approx(999_999.0)


def test_calculate_irr_npv_uses_session_cap_when_exit_zero(basic_session):
    from app.logic.calculator import calculate_proforma
    result = calculate_proforma(basic_session)
    basic_session.exit_cap_rate = 0.0
    data = calculate_irr_npv(basic_session, result)
    expected_exit = result["nois"][-1] / basic_session.cap_rate
    assert data["exit_value"] == pytest.approx(expected_exit)


def test_calculate_irr_npv_returns_all_keys(basic_session):
    from app.logic.calculator import calculate_proforma
    result = calculate_proforma(basic_session)
    data = calculate_irr_npv(basic_session, result)
    for key in ("irr", "npv", "exit_value", "effective_purchase", "effective_exit_cap"):
        assert key in data


def test_inputs_sheet_contains_return_analysis(basic_session, tmp_path):
    from openpyxl import load_workbook
    from app.logic.calculator import calculate_proforma, calculate_irr_npv
    from app.excel.writer import write_workbook
    result = calculate_proforma(basic_session)
    irr_data = calculate_irr_npv(basic_session, result)
    path = str(tmp_path / "test.xlsx")
    write_workbook(basic_session, path, irr_data=irr_data)
    wb = load_workbook(path)
    ws = wb["Inputs"]
    all_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any("Return Analysis" in str(v) for v in all_values if v)


def test_inputs_sheet_no_return_analysis_when_no_irr_data(basic_session, tmp_path):
    from openpyxl import load_workbook
    from app.excel.writer import write_workbook
    path = str(tmp_path / "test.xlsx")
    write_workbook(basic_session, path, irr_data=None)
    wb = load_workbook(path)
    ws = wb["Inputs"]
    all_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert not any("Return Analysis" in str(v) for v in all_values if v)
