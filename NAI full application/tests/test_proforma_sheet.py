from openpyxl import Workbook
from app.excel.inputs_sheet import write_inputs_sheet
from app.excel.proforma_sheet import write_proforma_sheet
from app.excel.cell_refs import PF_PROJ_START_COL, PF_TENANT_DATA_START


def _build(basic_session):
    wb = Workbook()
    ws_in = wb.active
    write_inputs_sheet(ws_in, basic_session)
    ws_pf = wb.create_sheet("ProForma")
    write_proforma_sheet(ws_pf, basic_session)
    return ws_pf


def test_tenant_year1_rent_is_IF_formula(basic_session):
    ws = _build(basic_session)
    cell = ws.cell(row=PF_TENANT_DATA_START, column=PF_PROJ_START_COL)
    assert str(cell.value).startswith("=IF(")


def test_sum_formula_exists(basic_session):
    ws = _build(basic_session)
    found = any(
        str(c.value).startswith("=SUM(")
        for row in ws.iter_rows() for c in row if c.value
    )
    assert found


def test_cap_rate_referenced(basic_session):
    ws = _build(basic_session)
    found = any(
        "Inputs!B11" in str(c.value)
        for row in ws.iter_rows() for c in row if c.value
    )
    assert found


def test_sensitivity_minus_delta(basic_session):
    ws = _build(basic_session)
    found = any(
        "Inputs!B11-Inputs!B12" in str(c.value).replace(" ", "")
        for row in ws.iter_rows() for c in row if c.value
    )
    assert found
